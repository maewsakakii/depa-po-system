from nicegui import ui
from datetime import datetime
from bahttext import bahttext
import io
import os
import json
import openpyxl
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# --- CONFIG ---
SHEET_NAME = "DEPA_PO_SYSTEM"
CURRENT_YEAR_TAB = "PO_2569"
TEMPLATE_FILE = "template_po.xlsx"
JSON_KEY_FILE = "service_account.json"

# --- STYLES ---
STYLE_INPUT = 'w-full'
PROPS_INPUT = 'outlined dense color="teal"'
STYLE_CARD = 'w-full max-w-6xl bg-white shadow-lg rounded-lg border border-gray-200 p-0 mx-auto'

# --- BACKEND: GOOGLE SHEET & LOGIC ---

def get_worksheet():
    """เชื่อมต่อ Google Sheet และสร้าง Header หากยังไม่มี"""
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    if not os.path.exists(JSON_KEY_FILE):
        ui.notify('ไม่พบไฟล์ service_account.json', type='negative')
        return None
    
    try:
        creds = ServiceAccountCredentials.from_json_keyfile_name(JSON_KEY_FILE, scope)
        client = gspread.authorize(creds)
        sheet = client.open(SHEET_NAME)
        try:
            ws = sheet.worksheet(CURRENT_YEAR_TAB)
        except:
            ws = sheet.add_worksheet(title=CURRENT_YEAR_TAB, rows=1000, cols=20)
            # Header Columns
            headers = [
                'PO No', 'Date', 'Project', 'PR No', 'Quote Info', 
                'Vendor Name', 'Tax ID', 'Grand Total', 'Preparer', 'Items_JSON'
            ]
            ws.append_row(headers)
        return ws
    except Exception as e:
        ui.notify(f"Connection Error: {e}", type='negative')
        return None

def get_next_po_number():
    """อ่าน Database เพื่อหาเลข PO ถัดไป"""
    ws = get_worksheet()
    if not ws: return "PO-69/001"
    
    try:
        # อ่านคอลัมน์ A (PO No) ทั้งหมด
        col_values = ws.col_values(1) # Column 1 = PO No
        if len(col_values) <= 1: # มีแค่ Header
            return "PO-69/001"
        
        last_po = col_values[-1] # เอาตัวล่าสุด
        # สมมติ format: PO-69/001 -> ตัดเอา 001 มา +1
        if "/" in last_po:
            prefix, running = last_po.split("/")
            next_num = int(running) + 1
            return f"{prefix}/{next_num:03d}"
        else:
            return "PO-69/001" # Format ผิด ให้เริ่มใหม่
    except:
        return "PO-69/001"

def fetch_po_history():
    """ดึงรายการ PO ทั้งหมดมาแสดงใน Dropdown"""
    ws = get_worksheet()
    if not ws: return {}
    
    # อ่านข้อมูลทั้งหมด (ข้าม Header)
    rows = ws.get_all_values()
    if len(rows) < 2: return {}
    
    # สร้าง Dict { 'PO-xxx': {data...} }
    history = {}
    for row in rows[1:]:
        # row index: 0=PO, 1=Date, 2=Project, ..., 9=Items_JSON
        if len(row) > 0:
            history[row[0]] = row
    return history

def save_to_database(state, grand_total):
    """บันทึกข้อมูลลง Sheet (ถ้ามีเลขเดิม = อัปเดต, ถ้าไม่มี = เพิ่มใหม่)"""
    ws = get_worksheet()
    if not ws: return False
    
    # แปลง Items เป็น JSON String เพื่อเก็บใน Cell เดียว
    items_json = json.dumps(state['items'], ensure_ascii=False)
    
    row_data = [
        state['po_no'], state['date'], state['project_name'], 
        state['pr_no'], f"{state['quote_no']} ({state['quote_date']})",
        state['vendor_name'], state['tax_id'], 
        f"{grand_total:.2f}", state['contact_person'], items_json
    ]

    try:
        # เช็คว่ามี PO นี้อยู่แล้วไหม
        cell = ws.find(state['po_no'], in_column=1)
        if cell:
            # Update Existing Row
            for col, val in enumerate(row_data, start=1):
                ws.update_cell(cell.row, col, val)
            ui.notify(f"อัปเดตข้อมูล {state['po_no']} เรียบร้อย", type='positive')
        else:
            # Append New Row
            ws.append_row(row_data)
            ui.notify(f"บันทึก PO ใหม่ {state['po_no']} เรียบร้อย", type='positive')
        return True
    except Exception as e:
        ui.notify(f"Save Error: {e}", type='negative')
        return False

# --- EXCEL GENERATION ---
def replace_text(ws, replacements):
    """ฟังก์ชันไล่แทนคำใน Excel"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for key, val in replacements.items():
                    target = f"{{{{ {key} }}}}" # {{ key }}
                    target_nospace = f"{{{{{key}}}}}" # {{key}}
                    if target in cell.value:
                        cell.value = cell.value.replace(target, str(val))
                    if target_nospace in cell.value:
                        cell.value = cell.value.replace(target_nospace, str(val))

def generate_excel(state, total_vars):
    if not os.path.exists(TEMPLATE_FILE):
        ui.notify(f"ไม่พบไฟล์ {TEMPLATE_FILE} กรุณาตรวจสอบโฟลเดอร์", type='negative')
        return None

    try:
        wb = openpyxl.load_workbook(TEMPLATE_FILE)
        ws = wb.active

        # 1. Prepare Data
        replacements = {
            'po_no': state['po_no'],
            'date': state['date'],
            'project_name': state['project_name'],
            'pr_no': state['pr_no'],
            'budget_code': state['budget_code'],
            'quote_no': state['quote_no'],
            'quote_date': state['quote_date'],
            'vendor_name': state['vendor_name'],
            'vendor_address': state['vendor_address'],
            'vendor_contact': state['vendor_contact'],
            'tax_id': state['tax_id'],
            'contact_person': state['contact_person'],
            'contact_ext': state['contact_ext'],
            'contact_email': state['contact_email'],
            'preparer': 'เจ้าหน้าที่พัสดุ',
            'subtotal': f"{total_vars['subtotal']:,.2f}",
            'vat_amount': f"{total_vars['vat']:,.2f}",
            'grand_total': f"{total_vars['grand_total']:,.2f}",
            'baht_text': bahttext(total_vars['grand_total'])
        }

        # 2. General Replace
        replace_text(ws, replacements)

        # 3. Item Table Logic (หาบรรทัดที่มี {{ item.desc }} แล้วเขียนทับ)
        start_row = 14 # Default fallback
        # ค้นหาบรรทัดเริ่ม
        found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'item.desc' in cell.value:
                    start_row = cell.row
                    found = True
                    break
            if found: break
        
        # เขียนรายการสินค้า
        current_row = start_row
        for i, item in enumerate(state['items']):
            line_total = float(item['qty']) * float(item['price'])
            # Mapping Column (ปรับแก้ A, B, H... ตามไฟล์จริงของคุณ)
            ws[f'A{current_row}'] = i + 1
            ws[f'B{current_row}'] = item['desc']
            ws[f'H{current_row}'] = float(item['qty'])
            ws[f'I{current_row}'] = item['unit']
            ws[f'J{current_row}'] = float(item['price'])
            ws[f'K{current_row}'] = line_total
            
            # Format
            ws[f'J{current_row}'].number_format = '#,##0.00'
            ws[f'K{current_row}'].number_format = '#,##0.00'
            current_row += 1

        # Clear remaining placeholders if items are few
        # (Optional: Clear rows below if needed)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    except Exception as e:
        ui.notify(f"Excel Error: {e}", type='negative')
        print(e)
        return None

# --- UI PAGE ---
@ui.page('/')
def main_page():
    ui.add_head_html("""
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Sarabun:wght@300;400;600&display=swap');
            body { font-family: 'Sarabun', sans-serif; background-color: #f3f4f6; }
        </style>
    """)

    # --- STATE ---
    # โหลดเลข PO ล่าสุดเมื่อเปิดโปรแกรม
    next_po = get_next_po_number()
    
    state = {
        'po_no': next_po,
        'date': datetime.now().strftime('%d/%m/%Y'),
        'project_name': '',
        'pr_no': '',
        'budget_code': '',
        'quote_no': '',
        'quote_date': '',
        'vendor_name': '',
        'vendor_address': '',
        'vendor_contact': '',
        'tax_id': '',
        'contact_person': 'พบธรรม',
        'contact_ext': '1131',
        'contact_email': 'pobthum.sa@depa.or.th',
        'items': [{'desc': '', 'qty': 1, 'unit': 'งาน', 'price': 0}],
    }

    # --- ACTIONS ---
    def calculate():
        total = sum(float(x['qty']) * float(x['price']) for x in state['items'])
        grand = total * 1.07
        label_grand.text = f"{grand:,.2f}"
        return grand

    def load_old_po(e):
        """โหลดข้อมูล PO เก่ามาใส่ Form"""
        po_id = e.value
        history = fetch_po_history()
        if po_id in history:
            row = history[po_id]
            # Map ข้อมูลจาก Row กลับเข้า State
            # row: [0=PO, 1=Date, 2=Project, 3=PR, 4=Quote, 5=Vendor, 6=Tax, 7=Total, 8=Prep, 9=ItemsJSON]
            state['po_no'] = row[0]
            state['date'] = row[1]
            state['project_name'] = row[2]
            state['pr_no'] = row[3]
            state['vendor_name'] = row[5]
            state['tax_id'] = row[6]
            
            # Parse Items JSON
            try:
                if len(row) >= 10:
                    state['items'] = json.loads(row[9])
                else:
                    ui.notify('ไม่พบข้อมูลสินค้าใน PO นี้ (Data Legacy)', type='warning')
            except:
                state['items'] = [{'desc': 'Error loading items', 'qty':0, 'price':0}]
            
            items_list.refresh()
            calculate()
            ui.notify(f'โหลดข้อมูล {po_id} แล้ว', type='positive')

    async def process_save_export():
        grand = calculate()
        total_vars = {'subtotal': grand/1.07, 'vat': grand - (grand/1.07), 'grand_total': grand}
        
        # 1. Save to DB
        success = save_to_database(state, grand)
        if not success: return

        # 2. Generate Excel
        excel_io = generate_excel(state, total_vars)
        if excel_io:
            filename = f"PO_{state['po_no'].replace('/', '-')}.xlsx"
            ui.download(excel_io.read(), filename)
            ui.notify('กำลังดาวน์โหลด...', type='positive')
        
        # 3. Refresh Next PO (กรณีสร้างใหม่)
        if state['po_no'] == next_po: 
            # ถ้าเป็นการสร้างใหม่ ให้เคลียร์ฟอร์มเตรียมรับเลขถัดไป
            pass # (หรือจะสั่งรีโหลดหน้าใหม่ก็ได้)

    # --- UI LAYOUT ---
    with ui.column().classes('w-full py-8 px-4 items-center'):
        
        # TOP BAR: History Loader
        with ui.card().classes('w-full max-w-6xl mb-4 bg-teal-50 border border-teal-200'):
            with ui.row().classes('w-full items-center justify-between p-2'):
                ui.label('🔄 ดึงข้อมูล PO เก่า / พิมพ์ซ้ำ').classes('font-bold text-teal-800')
                # Dropdown เลือก PO
                po_options = list(fetch_po_history().keys())
                ui.select(options=po_options, label='ค้นหาเลข PO', on_change=load_old_po).props('outlined dense options-dense use-input input-debounce="0"').classes('w-64 bg-white')

        with ui.card().classes(STYLE_CARD):
            # Header
            with ui.row().classes('w-full bg-teal-800 p-4 rounded-t-lg items-center justify-between'):
                ui.label('ระบบออกใบสั่งซื้อ (Auto-Run & Database)').classes('text-white text-xl font-bold')
                ui.button('รีเซ็ต / เลขใหม่', on_click=lambda: ui.open('/')).props('flat color=white icon=refresh')

            with ui.column().classes('p-6 w-full gap-4'):
                
                # SECTION 1: DOC INFO
                ui.label('ข้อมูลเอกสาร').classes(STYLE_LABEL)
                with ui.grid(columns=4).classes('w-full gap-4'):
                    ui.input('เลขที่ PO (Auto)').bind_value(state, 'po_no').props(PROPS_INPUT)
                    ui.input('วันที่').bind_value(state, 'date').props(PROPS_INPUT)
                    ui.input('อ้างอิงใบเสนอราคา').bind_value(state, 'quote_no').props(PROPS_INPUT)
                    ui.input('ลงวันที่ (Quote)').bind_value(state, 'quote_date').props(PROPS_INPUT)

                with ui.grid(columns=3).classes('w-full gap-4'):
                    ui.input('เลขที่ PR').bind_value(state, 'pr_no').props(PROPS_INPUT)
                    ui.input('รหัสงบประมาณ').bind_value(state, 'budget_code').props(PROPS_INPUT)
                    ui.input('ชื่องาน/โครงการ').bind_value(state, 'project_name').props(PROPS_INPUT)

                ui.separator()

                # SECTION 2: VENDOR
                ui.label('ข้อมูลผู้ขาย').classes(STYLE_LABEL)
                with ui.grid(columns=2).classes('w-full gap-4'):
                    ui.input('ชื่อผู้ขาย').bind_value(state, 'vendor_name').props(PROPS_INPUT)
                    ui.input('เลขผู้เสียภาษี').bind_value(state, 'tax_id').props(PROPS_INPUT)
                    ui.textarea('ที่อยู่').bind_value(state, 'vendor_address').props(PROPS_INPUT).classes('col-span-2')
                    ui.input('ผู้ติดต่อ (Vendor)').bind_value(state, 'vendor_contact').props(PROPS_INPUT).classes('col-span-2')

                ui.separator()

                # SECTION 3: ITEMS
                ui.label('รายการสินค้า').classes(STYLE_LABEL)
                @ui.refreshable
                def items_list():
                    with ui.row().classes('w-full gap-2 px-2'):
                        ui.label('รายการ').classes('flex-grow text-xs text-gray-500')
                        ui.label('จำนวน').classes('w-20 text-xs text-gray-500')
                        ui.label('หน่วย').classes('w-20 text-xs text-gray-500')
                        ui.label('ราคา/หน่วย').classes('w-28 text-xs text-gray-500')

                    for i, item in enumerate(state['items']):
                        with ui.row().classes('w-full gap-2 mb-1 items-start'):
                            ui.textarea().bind_value(item, 'desc').props('outlined dense rows=1').classes('flex-grow')
                            ui.number(on_change=calculate).bind_value(item, 'qty').props(PROPS_INPUT).classes('w-20')
                            ui.input().bind_value(item, 'unit').props(PROPS_INPUT).classes('w-20')
                            ui.number(on_change=calculate).bind_value(item, 'price').props(PROPS_INPUT).classes('w-28')
                            ui.button(icon='close', on_click=lambda idx=i: (state['items'].pop(idx), items_list.refresh(), calculate())).props('flat dense color=red round').classes('mt-1')
                            
                    ui.button('เพิ่มแถว', on_click=lambda: (state['items'].append({'desc':'', 'qty':1, 'price':0, 'unit':''}), items_list.refresh())).props('flat dense icon=add color=teal')
                items_list()

                # SECTION 4: FOOTER & ACTIONS
                with ui.row().classes('w-full justify-between items-end mt-6'):
                    # Internal Contact
                    with ui.column().classes('w-1/2 gap-2'):
                        ui.label('ผู้ประสานงาน (DEPA)').classes(STYLE_LABEL)
                        with ui.row().classes('w-full gap-2'):
                            ui.input('ชื่อ').bind_value(state, 'contact_person').props(PROPS_INPUT).classes('flex-grow')
                            ui.input('เบอร์ต่อ').bind_value(state, 'contact_ext').props(PROPS_INPUT).classes('w-24')
                    
                    # Totals
                    with ui.column().classes('items-end'):
                        ui.label('ยอดสุทธิ (รวม VAT)').classes('text-sm text-gray-600')
                        label_grand = ui.label('0.00').classes('text-3xl font-bold text-teal-800')

                ui.separator().classes('my-4')
                ui.button('💾 บันทึกและดาวน์โหลด Excel', on_click=process_save_export).props('unelevated color=teal icon=file_download w-full size=lg')

ui.run(title='DEPA PO System (DB Connected)', port=8080)
